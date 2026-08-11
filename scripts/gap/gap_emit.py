#!/usr/bin/env python3
"""Emit the amended xlsx and the two encrypted site payloads."""
import json, hashlib, gzip, os, secrets, re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

PASS = b'XMa5-5Bea-QgLV-aBvY'
OUT = '/tmp/build'
BASE = '/mnt/user-data/uploads/waves/2. Areas/Service/Office of Research/I&E/Landscape Analysis/UCSB GAP Landscape Database (2026-07-16).xlsx'
XLSX = f'{OUT}/UCSB GAP Landscape Database (2026-08-10).xlsx'
rows = json.load(open(f'{OUT}/data.json'))
N = len(rows)
LAST = N + 1                      # last data row in Directory

# column order: original 18, then the 5 amendment columns
KEYS = ['id', 'name', 'category', 'layer', 'status', 'status_detail', 'operator',
        'what_it_provides', 'award_or_value', 'sector_focus', 'eligibility', 'stage_served',
        'cadence', 'url', 'url_verified', 'institutional_participation', 'notes', 'sources',
        'resource_type', 'ucsb_contact', 'listing_consent', 'targets_spinouts', 'public_listing',
        'call_status']
HEAD = ['ID', 'Program / Vehicle', 'GAP Category', 'Layer', 'Status', 'Status Detail (evidence, dated)',
        'Operator', 'What It Provides', 'Award / Check Size', 'Sector Focus', 'Eligibility',
        'Stage Served', 'Cadence', 'URL', 'URL Verified', 'UCSB/UC Upside Share', 'Notes (strategic)',
        'Sources', 'Resource Type', 'UCSB Contact', 'Listing Consent', 'Targets University Spinouts',
        'Public Build', 'Call Status']

wb = openpyxl.load_workbook(BASE)
ws = wb['Directory']

# wipe every existing data row, then rewrite from the amended dataset
ws.auto_filter.ref = None
for r in range(ws.max_row, 1, -1):
    ws.delete_rows(r)
hdr_style = [ws.cell(row=1, column=c) for c in range(1, 19)]
for i, h in enumerate(HEAD, start=1):
    cell = ws.cell(row=1, column=i, value=h)
    if i > 18:
        src = hdr_style[0]
        cell.font = src.font.copy(); cell.fill = src.fill.copy()
        cell.border = src.border.copy(); cell.alignment = src.alignment.copy()
for ri, row in enumerate(rows, start=2):
    for ci, k in enumerate(KEYS, start=1):
        v = row.get(k, '')
        if isinstance(v, (list, tuple)):
            v = ' ; '.join(str(x) for x in v)
        elif isinstance(v, dict):
            v = json.dumps(v, ensure_ascii=False)
        ws.cell(row=ri, column=ci, value=(v if v not in ('', None) else None))
ws.auto_filter.ref = f'A1:{get_column_letter(len(HEAD))}{LAST}'
ws.freeze_panes = 'B2'
for ci in range(19, len(HEAD) + 1):
    ws.column_dimensions[get_column_letter(ci)].width = 26

# Summary: extend every COUNTIF/COUNTIFS range from row 84 to the new last row
sm = wb['Summary']
for row in sm.iter_rows():
    for c in row:
        if isinstance(c.value, str) and c.value.startswith('='):
            c.value = re.sub(r'(\$[A-Z]\$2:\$[A-Z]\$)84', rf'\g<1>{LAST}', c.value)
sm.cell(row=21, column=2,
        value=(f'Source: 2026-07-16 master landscape research pass, amended {"2026-08-10"} from the '
               f'2026-07 reviewer round (see "GAP Landscape — Feedback Amendments (July 2026)"). '
               f'{N} entries. Counts recalculate automatically if Directory rows are edited or added '
               f'within the referenced range.'))

# README: restate compiled line, counts and the POC key finding
rd = wb['README']
for row in rd.iter_rows():
    for c in row:
        v = c.value
        if not isinstance(v, str):
            continue
        if v.startswith('Compiled'):
            c.value = ('Compiled 2026-07-16 by Kelly Caylor (AVC Research / Innovation) with Claude '
                       'research assistance. Amended 2026-08-10 from the July 2026 reviewer round '
                       '(Englander, Margalith, Gartner, Driscoll, Cotter).')
        elif v.startswith('Master table'):
            c.value = (f'Master table - one row per program, facility, event or service ({N} entries). '
                       'Auto-filter on; slice by GAP category, resource type, layer or status.')
        elif v.startswith('Weakest layer'):
            c.value = ('REVISED 2026-08-10. The UC systemwide POC program was renewed on 2026-07-29: '
                       'UCSB sits in the no-match tier at up to $200,000/yr for three years '
                       '(FY 2026-27 to 2028-29), first awards possible 2026-10-01. ACTIVATE remains '
                       'suspended pending confirmation that it is the local delivery vehicle. '
                       'CNSI SEED-TECH is active with no open call. Vandenberg POC (~$262K/yr) opens '
                       'its first call in Fall 2026. The layer is materially stronger than the '
                       '2026-07-16 assessment stated.')
        elif v.startswith('Local early-stage check-writers'):
            c.value = ('Local early-stage check-writers: SBAA, ScOp, Entrada, Cycad, AngelCon '
                       '(~$155K/yr), plus Central Coast Ventures (SLO) added 2026-08-10. TCA '
                       'coverage is now recorded as Pasadena/LA rather than a standing Santa Barbara '
                       'chapter, pending confirmation. SBVP is Series A-C; Rincon\'s franchise moved '
                       'to LA (Bonfire).')
        elif v.startswith('Entries with url_verified'):
            c.value = ('Entries with URL Verified = FALSE or Status = Unclear need direct confirmation '
                       'before public listing. The Public Build column carries the filter decision for '
                       'the shared version: defunct entries are hidden, and investors without recorded '
                       'listing consent are hidden. Internal-only figures (Vandenberg $525K POC pool) '
                       'are from TIA briefings, not public sources.')
wb.save(XLSX)

# ------------------------------------------------------------- encrypt
def seal(plain: bytes) -> bytes:
    salt, iv = secrets.token_bytes(16), secrets.token_bytes(12)
    k = hashlib.pbkdf2_hmac('sha256', PASS, salt, 310000, 32)
    return salt + iv + AESGCM(k).encrypt(iv, plain, None)

def unseal(blob: bytes) -> bytes:
    k = hashlib.pbkdf2_hmac('sha256', PASS, blob[:16], 310000, 32)
    return AESGCM(k).decrypt(blob[16:28], blob[28:], None)

data_plain = gzip.compress(json.dumps(rows, ensure_ascii=False, separators=(',', ':')).encode())
open(f'{OUT}/data.enc', 'wb').write(seal(data_plain))
open(f'{OUT}/xlsx.enc', 'wb').write(seal(open(XLSX, 'rb').read()))

# round-trip verification against the exact code path the browser uses
d = json.loads(gzip.decompress(unseal(open(f'{OUT}/data.enc', 'rb').read())))
x = unseal(open(f'{OUT}/xlsx.enc', 'rb').read())
assert len(d) == N and d[0]['name'] and x[:2] == b'PK'
wb2 = openpyxl.load_workbook(f'{OUT}/{os.path.basename(XLSX)}')
assert wb2['Directory'].max_row == LAST, wb2['Directory'].max_row
print(f'xlsx      {os.path.getsize(XLSX):>8,} bytes  Directory rows={wb2["Directory"].max_row - 1} cols={wb2["Directory"].max_column}')
print(f'data.enc  {os.path.getsize(OUT + "/data.enc"):>8,} bytes  round-trip {len(d)} rows OK')
print(f'xlsx.enc  {os.path.getsize(OUT + "/xlsx.enc"):>8,} bytes  round-trip PK header OK')
